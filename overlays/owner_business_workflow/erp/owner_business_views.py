from __future__ import annotations

import csv
import hashlib
import io
import json
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.views.decorators.http import require_GET, require_POST

from . import models as m
from .rebuild_views import _is_field_user, _org
from .services.org_price_search import search_org_prices, serialize_org_price


MAX_UPLOAD_BYTES = 20 * 1024 * 1024
MAX_ROWS = 20000


def _can_manage_price_lists(request) -> bool:
    if not getattr(request.user, "is_authenticated", False):
        return False
    if getattr(request.user, "is_superuser", False) or getattr(request.user, "is_staff", False):
        return True
    if _is_field_user(request):
        return False
    profile = getattr(request.user, "profile", None)
    role = str(getattr(profile, "role", "office") or "office").casefold()
    return role in {"owner", "admin", "manager", "office", "backoffice", "büro", "buero"}


def _norm_header(value: str) -> str:
    value = str(value or "").strip().casefold()
    value = value.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    return re.sub(r"[^a-z0-9]+", "", value)


HEADER_ALIASES = {
    "code": {"code", "artikelnummer", "artikelnr", "artnr", "positionsnummer", "positionsnr", "position", "nummer", "nr", "sku", "va04"},
    "description": {"beschreibung", "bezeichnung", "leistung", "artikel", "text", "langtext", "description", "name"},
    "unit": {"einheit", "mengeneinheit", "unit", "me"},
    "sales_price": {"verkaufspreis", "vk", "vkpreis", "preis", "einzelpreis", "nettopreis", "salesprice", "unitprice"},
    "purchase_price": {"einkaufspreis", "ek", "ekpreis", "purchaseprice", "cost", "kosten"},
    "tax_rate": {"mwst", "mwstsatz", "steuer", "steuersatz", "tax", "taxrate"},
}


def _decimal(value) -> Decimal | None:
    if value is None:
        return None
    raw = str(value).strip().replace("€", "").replace("EUR", "").replace("eur", "").replace(" ", "")
    if not raw:
        return None
    if raw.count(",") == 1 and raw.count(".") >= 1 and raw.rfind(",") > raw.rfind("."):
        raw = raw.replace(".", "").replace(",", ".")
    elif raw.count(",") == 1 and "." not in raw:
        raw = raw.replace(",", ".")
    try:
        return Decimal(raw)
    except (InvalidOperation, ValueError):
        return None


def _rows_from_csv(raw: bytes) -> list[list[str]]:
    text = raw.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ";" if sample.count(";") >= sample.count(",") else ","
    return [list(row) for row in csv.reader(io.StringIO(text), delimiter=delimiter)]


def _rows_from_xlsx(raw: bytes) -> list[list[object]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("XLSX-Import ist auf diesem Server nicht verfügbar. Bitte CSV verwenden.") from exc
    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheet = workbook.active
    rows = []
    for index, row in enumerate(sheet.iter_rows(values_only=True)):
        rows.append(list(row))
        if index >= MAX_ROWS:
            break
    workbook.close()
    return rows


def _read_rows(filename: str, raw: bytes) -> list[list[object]]:
    suffix = Path(filename or "").suffix.casefold()
    if suffix in {".xlsx", ".xlsm"}:
        return _rows_from_xlsx(raw)
    if suffix in {".csv", ".txt", ".tsv"}:
        return _rows_from_csv(raw)
    raise ValueError("Bitte eine CSV-, TSV- oder XLSX-Datei hochladen.")


def _detect_columns(header: list[object]) -> dict[str, int]:
    normalized = [_norm_header(value) for value in header]
    result: dict[str, int] = {}
    for key, aliases in HEADER_ALIASES.items():
        alias_norm = {_norm_header(alias) for alias in aliases}
        for index, value in enumerate(normalized):
            if value in alias_norm:
                result[key] = index
                break
    if "description" not in result:
        raise ValueError("Keine Spalte für Beschreibung/Bezeichnung gefunden.")
    if "sales_price" not in result and "purchase_price" not in result:
        raise ValueError("Keine Preis-Spalte gefunden. Erwartet z. B. Verkaufspreis/VK/Preis oder Einkaufspreis/EK.")
    return result


def _cell(row: list[object], columns: dict[str, int], key: str):
    index = columns.get(key)
    return row[index] if index is not None and index < len(row) else None


def _model_kwargs(model, values: dict) -> dict:
    fields = {field.name for field in model._meta.get_fields()}
    return {key: value for key, value in values.items() if key in fields}


def _import_price_rows(*, org, user, name: str, filename: str, raw: bytes, replace_previous: bool) -> tuple[object, int, int]:
    rows = _read_rows(filename, raw)
    rows = [row for row in rows if any(str(cell or "").strip() for cell in row)]
    if len(rows) < 2:
        raise ValueError("Die Preisdatei enthält keine importierbaren Datenzeilen.")
    columns = _detect_columns(rows[0])
    digest = hashlib.sha256(raw).hexdigest()

    if replace_previous:
        m.PriceSource.objects.filter(organization=org, name__iexact=name, active=True).update(active=False)

    source_values = {
        "organization": org,
        "name": name[:200],
        "original_filename": (filename or "preisliste")[:255],
        "sha256": digest,
        "active": True,
        "import_summary": {"kind": "owner_upload", "uploaded_by": getattr(user, "pk", None), "columns": columns},
    }
    source = m.PriceSource.objects.create(**_model_kwargs(m.PriceSource, source_values))
    price_fields = {field.name for field in m.PriceItem._meta.get_fields()}
    objects = []
    skipped = 0
    for row_index, row in enumerate(rows[1:MAX_ROWS + 1], start=1):
        description = str(_cell(row, columns, "description") or "").strip()
        if not description:
            skipped += 1
            continue
        sales = _decimal(_cell(row, columns, "sales_price"))
        purchase = _decimal(_cell(row, columns, "purchase_price"))
        if sales is None and purchase is None:
            skipped += 1
            continue
        code = str(_cell(row, columns, "code") or "").strip() or f"IMP-{source.pk}-{row_index:05d}"
        unit = str(_cell(row, columns, "unit") or "Stk.").strip() or "Stk."
        values = {
            "organization": org,
            "source": source,
            "code": code[:120],
            "description": description[:1000],
            "unit": unit[:50],
            "sales_price": sales,
            "purchase_price": purchase,
            "tax_rate": _decimal(_cell(row, columns, "tax_rate")),
        }
        values = {key: value for key, value in values.items() if key in price_fields and value is not None}
        objects.append(m.PriceItem(**values))
    if not objects:
        source.delete()
        raise ValueError("Es wurde keine Zeile mit Beschreibung und gültigem Preis gefunden.")
    m.PriceItem.objects.bulk_create(objects, batch_size=1000)
    summary = getattr(source, "import_summary", None)
    if isinstance(summary, dict):
        summary.update({"imported": len(objects), "skipped": skipped, "sha256": digest})
        source.import_summary = summary
        source.save(update_fields=["import_summary"])
    return source, len(objects), skipped


@login_required
@require_POST
def price_list_upload(request):
    if not _can_manage_price_lists(request):
        return JsonResponse({"ok": False, "error": "Keine Berechtigung für Preislisten."}, status=403)
    org = _org(request)
    upload = request.FILES.get("price_file")
    if upload is None:
        messages.error(request, "Bitte eine Preisdatei auswählen.")
        return redirect("next-settings")
    if getattr(upload, "size", 0) > MAX_UPLOAD_BYTES:
        messages.error(request, "Die Preisdatei darf maximal 20 MB groß sein.")
        return redirect("next-settings")
    name = (request.POST.get("price_list_name") or Path(upload.name).stem or "Eigene Preisliste").strip()
    if not name:
        name = "Eigene Preisliste"
    try:
        source, imported, skipped = _import_price_rows(
            org=org,
            user=request.user,
            name=name,
            filename=upload.name,
            raw=upload.read(),
            replace_previous=request.POST.get("replace_previous") == "1",
        )
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect("next-settings")
    messages.success(request, f"Preisliste „{source.name}“ importiert: {imported} Positionen" + (f", {skipped} übersprungen." if skipped else "."))
    return redirect("next-settings")


@login_required
@require_POST
def price_list_toggle(request, pk):
    if not _can_manage_price_lists(request):
        return JsonResponse({"ok": False, "error": "Keine Berechtigung für Preislisten."}, status=403)
    org = _org(request)
    source = get_object_or_404(m.PriceSource, pk=pk, organization=org)
    source.active = not bool(source.active)
    source.save(update_fields=["active"])
    messages.success(request, f"Preisliste „{source.name}“ ist jetzt {'aktiv' if source.active else 'deaktiviert'}.")
    return redirect("next-settings")


@login_required
@require_GET
def organization_price_search(request):
    if _is_field_user(request):
        return JsonResponse({"ok": False, "error": "Keine Preisberechtigung."}, status=403)
    org = _org(request)
    query = (request.GET.get("q") or "").strip()
    results = [serialize_org_price(row) for row in search_org_prices(org, query, limit=30)]
    return JsonResponse({"ok": True, "results": results})
